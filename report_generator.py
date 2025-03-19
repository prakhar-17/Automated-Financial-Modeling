# report_generator.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def generate_report(ticker, forecast_df, dcf_value):
    """
    Generates an Excel report with forecasts, valuation, and charts.
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    
    # Add forecast data
    ws.title = "Forecast"
    for row in pd.DataFrame(forecast_df).itertuples(index=False):
        ws.append(row)
    
    # Add valuation data
    ws = wb.create_sheet("Valuation")
    ws.append(["DCF Valuation", dcf_value])
    
    # Add charts
    ws = wb.create_sheet("Charts")
    img = Image("examples/AAPL_charts/forecast.png")
    ws.add_image(img, "A1")
    
    img = Image("examples/AAPL_charts/revenue_growth.png")
    ws.add_image(img, "A20")
    
    img = Image("examples/AAPL_charts/profit_margins.png")
    ws.add_image(img, "A40")
    
    img = Image("examples/AAPL_charts/stock_prices.png")
    ws.add_image(img, "A60")
    
    img = Image("examples/AAPL_charts/sensitivity_analysis.png")
    ws.add_image(img, "A80")
    
    # Save the workbook
    wb.save(f"{ticker}_valuation_report.xlsx")
    print(f"Report saved to {ticker}_valuation_report.xlsx")